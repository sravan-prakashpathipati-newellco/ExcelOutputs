# import the Flask class from the flask module
from flask import Flask, render_template,request,send_from_directory,send_file,Response
import flask_excel as excel
import openpyxl
import csv
import urllib.request as urllib2
import openpyxl
import io
from io import BytesIO
from io import StringIO
from urllib.parse import quote
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.drawing.image import Image
import PIL
from PIL import Image as immmg
import io
import urllib3
from collections import defaultdict

try:
    from urllib.request import urlopen
except ImportError:
    from urllib2 import urlopen

# create the application object
app = Flask(__name__)
app.config["USE_X_SENDFILE"] = True

@app.route('/page')
def welcome():
    return render_template('page.html')  # render a template


# use decorators to link the function to a url
@app.route('/parse',methods=["POST"])
def home():

    if request.method == "POST":
        if request.files:
            #import pdb; pdb.set_trace()
            file = request.files['file']
            file_name = file.filename
            all_rows= []
            output_file_name = file_name.split('.')[0] +'_output'+ '.csv'


            wb = openpyxl.load_workbook(file)

            sheet = wb.active

            column_names=[]
            for cell in sheet[1]:
                column_names.append(cell.value)

            all_rows.append(column_names)


            id = sheet['A'][1:]
            name = sheet['B'][1:]
            prod_to_image = sheet['C'][1:]
            prod_to_image_name = sheet['D'][1:]
            dash = sheet['E'][1:]
            sort = sheet['F'][1:]
            s7_url = sheet['G'][1:]
            #asset_types = sheet['H'][1:]
            #alt_texts = sheet['I'][1:]




            for i,val in enumerate(id):
                #import pdb; pdb.set_trace()
                value = str(val.value)

                prod_image_val = str(prod_to_image[i].value).split(';')
                prod_image_names = str(prod_to_image_name[i].value).split(';')
                dash_val = str(dash[i].value).split(';')
                sort_val = str(sort[i].value).split(';')
                s7_val = str(s7_url[i].value).split(';')
                #asset_type = str(asset_types[i].value).split(';')
                #alt_text = str(alt_texts[i].value).split(';')




                for j,each in enumerate(prod_image_val):

                    to_write = []
                    to_write.append(value)
                    to_write.append(str(name[i].value))

                    to_write.append(each)
                    to_write.append(prod_image_names[j])

                    to_write.append(sort_val[j])
                    to_write.append(dash_val[j])
                    to_write.append(s7_val[j])
                    #to_write.append(asset_type[j])
                    #to_write.append(alt_text[j])




                    all_rows.append(to_write)

        return excel.make_response_from_array(all_rows, "csv",
                                      file_name=output_file_name)

@app.route('/getUrls',methods=["POST"])
def getUrls():

    if request.method == "POST":
        if request.files:

            file = request.files['file']
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            dict = []

            urls = sheet['B'][1:]
            name = sheet['A'][1:]

            for i,val in enumerate(urls):

                dict.append([name[i].value,urls[i].value,i+2])
            wb.close()

            import xlsxwriter
            # Create the workbook and add a worksheet.

            output = StringIO()
            workbook  = xlsxwriter.Workbook('output',{'in_memory': True})
            worksheet = workbook.add_worksheet()

            urls = dict

            worksheet.set_column(2, 2, 50)
            worksheet.set_column(3, 3, 25)
            worksheet.set_default_row(50)
            text_format = workbook.add_format({'text_wrap': True})

            #worksheet.write('A1', 'Name',text_format)
            #worksheet.write('B1', 'URL',text_format)
            #worksheet.write('C1', 'Image Preview',text_format)

            for name,url,value in urls:

                image_data = BytesIO(urlopen(url).read())

                # Write the byte stream image to a cell. Note, the filename must be
                # specified. In this case it will be read from url string.
                cell = 'C{}'.format(value)
                worksheet.write('A{}'.format(value), name,text_format)
                worksheet.write('B{}'.format(value), url,text_format)
                worksheet.insert_image(cell, url, {'image_data': image_data,'x_scale': 0.20, 'y_scale': 0.18})

            #worksheet.seek(0)

            #worksheet.set_column('C', my_format)
            #worksheet.seek(0)
            workbook.close()
            output.seek(0)
            return send_file(
                            workbook,
                            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            as_attachment=True,
                            attachment_filename="output.xlsx"
                        )


# start the server with the 'run()' method

if __name__ == '__main__':
    excel.init_excel(app)
    app.run(debug=True)
